"""Generate an Excel template for candidates upload.

Output: files/candidates_template.xlsx

The phone column is formatted as text so Excel does not turn '+79151234567'
into a number / scientific notation. Includes one example row.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

OUT = Path(__file__).resolve().parents[1] / "files" / "candidates_template.xlsx"


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "candidates"

    headers = ["Телефон", "ФИО", "Источник"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")

    for col_idx in range(1, len(headers) + 1):
        for row_idx in range(1, 1001):
            ws.cell(row=row_idx, column=col_idx).number_format = "@"

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 16

    ws.append(["+79151234567", "Иванов Иван Иванович", "hh.ru"])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
