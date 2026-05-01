"""Excel-выгрузка кандидатов вакансии для HR.

Один файл = одна вакансия = один лист «Кандидаты». Каждая строка —
кандидат + его последний звонок (для аналитики глубже есть UI).

Используем openpyxl (он уже подключён для импорта /candidates/upload),
без pandas/xlsxwriter.
"""

from __future__ import annotations

import io
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_MSK = ZoneInfo("Europe/Moscow")

_HEADERS = [
    ("ФИО", 28),
    ("Телефон", 16),
    ("Источник", 14),
    ("Статус", 14),
    ("Попыток", 9),
    ("Дата последнего звонка (МСК)", 22),
    ("Длительность, сек", 17),
    ("Score", 8),
    ("Decision", 12),
    ("Reasoning", 50),
    ("Резюме", 60),
    ("Ссылка на запись", 50),
    ("Карточка звонка", 50),
]

_DECISION_FILL = {
    "pass": PatternFill("solid", fgColor="C6EFCE"),
    "reject": PatternFill("solid", fgColor="FFC7CE"),
    "review": PatternFill("solid", fgColor="FFEB9C"),
    "not_reached": PatternFill("solid", fgColor="EAEAEA"),
}
_DECISION_LABEL = {
    "pass": "pass",
    "reject": "reject",
    "review": "review",
    "not_reached": "not_reached",
}


def _fmt_dt_msk(dt: datetime | None) -> str:
    if dt is None:
        return ""
    # БД хранит UTC naive — досаживаем UTC и переводим в МСК.
    return dt.replace(tzinfo=ZoneInfo("UTC")).astimezone(_MSK).strftime("%Y-%m-%d %H:%M")


def build_candidates_xlsx(rows: list[dict]) -> bytes:
    """Собрать xlsx-файл из списка строк.

    Каждая строка должна содержать ключи:
      fio, phone, source, status, attempts_count,
      last_started_at (datetime | None), last_duration (int | None),
      last_score (float | None), last_decision (str | None),
      last_reasoning (str | None), last_summary (str | None),
      recording_url (str | None), call_card_url (str | None).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Кандидаты"

    # Шапка
    header_font = Font(bold=True)
    for col_idx, (label, width) in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    # Строки
    for row_idx, r in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=r.get("fio") or "")
        ws.cell(row=row_idx, column=2, value=r.get("phone") or "")
        ws.cell(row=row_idx, column=3, value=r.get("source") or "")
        ws.cell(row=row_idx, column=4, value=r.get("status") or "")
        ws.cell(row=row_idx, column=5, value=r.get("attempts_count") or 0)
        ws.cell(row=row_idx, column=6, value=_fmt_dt_msk(r.get("last_started_at")))
        ws.cell(row=row_idx, column=7, value=r.get("last_duration") or "")
        score = r.get("last_score")
        ws.cell(row=row_idx, column=8, value=round(score, 1) if isinstance(score, (int, float)) else "")
        decision = r.get("last_decision") or ""
        d_cell = ws.cell(row=row_idx, column=9, value=_DECISION_LABEL.get(decision, decision))
        if decision in _DECISION_FILL:
            d_cell.fill = _DECISION_FILL[decision]
        ws.cell(row=row_idx, column=10, value=r.get("last_reasoning") or "")
        ws.cell(row=row_idx, column=11, value=r.get("last_summary") or "")
        ws.cell(row=row_idx, column=12, value=r.get("recording_url") or "")
        ws.cell(row=row_idx, column=13, value=r.get("call_card_url") or "")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
