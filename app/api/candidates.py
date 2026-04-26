"""Candidate ingestion: bulk upload from XLSX/CSV + per-candidate lookup."""

from __future__ import annotations

import csv
import io
import re

import structlog
from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import desc, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_principal
from app.db.models import Call, Candidate, Client, Vacancy
from app.db.session import get_session
from app.workers.tasks import initiate_call

router = APIRouter()
log = structlog.get_logger()


PHONE_HEADERS = {"phone", "телефон", "тел", "номер", "mobile"}
NAME_HEADERS = {"fio", "name", "фио", "имя"}
SOURCE_HEADERS = {"source", "источник"}


def normalize_phone(raw: object) -> str | None:
    """Normalize a phone value (str or numeric) to E.164 +7XXXXXXXXXX. Returns None if invalid."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    digits = re.sub(r"\D", "", s)
    if len(digits) == 11 and digits[0] in {"7", "8"}:
        digits = "7" + digits[1:]
    elif len(digits) == 10:
        digits = "7" + digits
    else:
        return None
    return "+" + digits


def _normalize_header(h: object) -> str:
    return str(h or "").strip().lower()


def _pick_columns(headers: list[str]) -> tuple[int | None, int | None, int | None]:
    """Return (phone_idx, name_idx, source_idx) by matching header names."""
    phone_idx = name_idx = source_idx = None
    for i, h in enumerate(headers):
        nh = _normalize_header(h)
        if phone_idx is None and nh in PHONE_HEADERS:
            phone_idx = i
        elif name_idx is None and nh in NAME_HEADERS:
            name_idx = i
        elif source_idx is None and nh in SOURCE_HEADERS:
            source_idx = i
    return phone_idx, name_idx, source_idx


def _parse_xlsx(data: bytes) -> list[list[object]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True) if any(c is not None for c in row)]


def _parse_csv(data: bytes) -> list[list[object]]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    return [row for row in reader if any((c or "").strip() for c in row)]


@router.post("/upload")
async def upload_candidates(
    vacancy_id: int = Query(..., description="Vacancy to attach candidates to"),
    start: bool = Query(False, description="Enqueue calls immediately for new candidates"),
    file: UploadFile = File(...),
    client: Client = Depends(get_current_principal),
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Upload candidates from XLSX or CSV. Dedup by (vacancy_id, phone).

    Required columns (header row, case-insensitive, RU/EN):
      - phone / телефон
      - fio   / фио
    Optional:
      - source / источник
    """
    vacancy = await session.get(Vacancy, vacancy_id)
    if vacancy is None or vacancy.client_id != client.id:
        raise HTTPException(status_code=404, detail="vacancy not found")

    raw = await file.read()
    name = (file.filename or "").lower()
    try:
        if name.endswith(".xlsx"):
            rows = _parse_xlsx(raw)
        elif name.endswith(".csv"):
            rows = _parse_csv(raw)
        else:
            raise HTTPException(status_code=400, detail="unsupported file type (use .xlsx or .csv)")
    except HTTPException:
        raise
    except Exception as exc:
        log.exception("upload_parse_failed", filename=file.filename, error=str(exc))
        raise HTTPException(status_code=400, detail=f"failed to parse file: {exc}") from exc

    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="file is empty or missing data rows")

    headers = [_normalize_header(h) for h in rows[0]]
    phone_idx, name_idx, source_idx = _pick_columns(headers)
    if phone_idx is None or name_idx is None:
        raise HTTPException(
            status_code=400,
            detail="missing required columns: phone/телефон and fio/фио",
        )

    existing_q = await session.execute(
        select(Candidate.phone).where(Candidate.vacancy_id == vacancy_id)
    )
    existing_phones = {p for (p,) in existing_q}

    seen_in_file: set[str] = set()
    created: list[Candidate] = []
    invalid: list[dict] = []
    duplicates = 0

    for row_num, row in enumerate(rows[1:], start=2):
        phone_raw = row[phone_idx] if phone_idx < len(row) else None
        fio_raw = row[name_idx] if name_idx < len(row) else None
        source_raw = (
            row[source_idx]
            if source_idx is not None and source_idx < len(row)
            else None
        )
        phone = normalize_phone(phone_raw)
        fio = str(fio_raw or "").strip()
        if not phone:
            invalid.append({"row": row_num, "reason": "invalid phone", "value": str(phone_raw)})
            continue
        if not fio:
            invalid.append({"row": row_num, "reason": "missing fio"})
            continue
        if phone in existing_phones or phone in seen_in_file:
            duplicates += 1
            continue
        seen_in_file.add(phone)
        created.append(
            Candidate(
                vacancy_id=vacancy_id,
                phone=phone,
                fio=fio,
                source=str(source_raw).strip() if source_raw else None,
            )
        )

    if created:
        session.add_all(created)
        await session.commit()
        for c in created:
            await session.refresh(c)

    enqueued = 0
    if start and created:
        from app.workers.tasks import initiate_call

        for c in created:
            initiate_call.delay(c.id)
            enqueued += 1

    log.info(
        "candidates_uploaded",
        client_id=client.id,
        vacancy_id=vacancy_id,
        created=len(created),
        duplicates=duplicates,
        invalid=len(invalid),
        enqueued=enqueued,
    )
    return {
        "vacancy_id": vacancy_id,
        "created": len(created),
        "duplicates": duplicates,
        "invalid": invalid,
        "enqueued": enqueued,
    }


@router.get("")
async def list_candidates(
    vacancy_id: int = Query(...),
    limit: int = Query(default=200, ge=1, le=1000),
    include_archived: bool = Query(default=False),
    client: Client = Depends(get_current_principal),
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Список кандидатов вакансии с агрегатом по последнему звонку."""
    vacancy = await session.get(Vacancy, vacancy_id)
    if vacancy is None or vacancy.client_id != client.id:
        raise HTTPException(status_code=404, detail="vacancy not found")

    # Берём кандидатов и LEFT JOIN на «последний звонок» (по max id).
    last_call_subq = (
        select(
            Call.candidate_id.label("cid"),
            func.max(Call.id).label("max_call_id"),
        )
        .group_by(Call.candidate_id)
        .subquery()
    )

    stmt = (
        select(Candidate, Call)
        .outerjoin(last_call_subq, last_call_subq.c.cid == Candidate.id)
        .outerjoin(Call, Call.id == last_call_subq.c.max_call_id)
        .where(Candidate.vacancy_id == vacancy_id)
    )
    if not include_archived:
        stmt = stmt.where(Candidate.active.is_(True))
    stmt = stmt.order_by(desc(Candidate.id)).limit(limit)
    rows = (await session.execute(stmt)).all()

    items = []
    for cand, last_call in rows:
        items.append(
            {
                "id": cand.id,
                "vacancy_id": cand.vacancy_id,
                "fio": cand.fio,
                "phone": cand.phone,
                "source": cand.source,
                "status": cand.status,
                "active": cand.active,
                "created_at": cand.created_at.isoformat(),
                "last_call": {
                    "id": last_call.id,
                    "started_at": last_call.started_at.isoformat() if last_call.started_at else None,
                    "score": last_call.score,
                    "decision": last_call.decision,
                }
                if last_call is not None
                else None,
            }
        )
    return {"items": items, "vacancy_id": vacancy_id}


@router.post("/{candidate_id}/call", status_code=202)
async def call_candidate(
    candidate_id: int,
    client: Client = Depends(get_current_principal),
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Поставить кандидату задачу на обзвон через Celery."""
    stmt = (
        select(Candidate)
        .join(Vacancy, Candidate.vacancy_id == Vacancy.id)
        .where(Candidate.id == candidate_id, Vacancy.client_id == client.id)
    )
    candidate = (await session.execute(stmt)).scalar_one_or_none()
    if candidate is None:
        raise HTTPException(status_code=404, detail="candidate not found")
    if not candidate.active:
        raise HTTPException(
            status_code=400, detail="кандидат архивирован — восстановите его перед обзвоном"
        )

    task = initiate_call.delay(candidate.id)
    log.info(
        "candidate_call_enqueued",
        client_id=client.id,
        candidate_id=candidate.id,
        task_id=task.id,
    )
    return {"candidate_id": candidate.id, "task_id": task.id}


class CandidateUpdate(BaseModel):
    fio: str | None = Field(default=None, min_length=1, max_length=255)
    phone: str | None = Field(default=None, min_length=4, max_length=20)
    source: str | None = Field(default=None, max_length=50)
    active: bool | None = None


@router.patch("/{candidate_id}")
async def update_candidate(
    candidate_id: int,
    payload: CandidateUpdate,
    client: Client = Depends(get_current_principal),
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Обновить ФИО / телефон / источник / статус активности кандидата."""
    stmt = (
        select(Candidate)
        .join(Vacancy, Candidate.vacancy_id == Vacancy.id)
        .where(Candidate.id == candidate_id, Vacancy.client_id == client.id)
    )
    candidate = (await session.execute(stmt)).scalar_one_or_none()
    if candidate is None:
        raise HTTPException(status_code=404, detail="candidate not found")

    changes = payload.model_dump(exclude_unset=True)
    if not changes:
        raise HTTPException(status_code=400, detail="нет полей для обновления")

    if "fio" in changes:
        candidate.fio = changes["fio"].strip()
    if "phone" in changes:
        candidate.phone = changes["phone"].strip()
    if "source" in changes:
        candidate.source = changes["source"].strip() if changes["source"] else None
    if "active" in changes:
        candidate.active = changes["active"]

    await session.commit()
    await session.refresh(candidate)

    log.info(
        "candidate_updated",
        client_id=client.id,
        candidate_id=candidate.id,
        fields=list(changes.keys()),
    )
    return {
        "id": candidate.id,
        "vacancy_id": candidate.vacancy_id,
        "fio": candidate.fio,
        "phone": candidate.phone,
        "source": candidate.source,
        "status": candidate.status,
        "active": candidate.active,
    }


@router.delete("/{candidate_id}", status_code=204)
async def archive_candidate(
    candidate_id: int,
    client: Client = Depends(get_current_principal),
    session: AsyncSession = Depends(get_session),
) -> Response:
    """Soft-архивирование. Идемпотентно. Записи звонков остаются."""
    stmt = (
        select(Candidate)
        .join(Vacancy, Candidate.vacancy_id == Vacancy.id)
        .where(Candidate.id == candidate_id, Vacancy.client_id == client.id)
    )
    candidate = (await session.execute(stmt)).scalar_one_or_none()
    if candidate is None:
        raise HTTPException(status_code=404, detail="candidate not found")

    if candidate.active:
        candidate.active = False
        await session.commit()
        log.info(
            "candidate_archived",
            client_id=client.id,
            candidate_id=candidate.id,
        )

    return Response(status_code=204)


@router.get("/{candidate_id}")
async def get_candidate(
    candidate_id: int,
    client: Client = Depends(get_current_principal),
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Get candidate details with call results, scoped to the calling client."""
    stmt = (
        select(Candidate)
        .join(Vacancy, Candidate.vacancy_id == Vacancy.id)
        .where(Candidate.id == candidate_id, Vacancy.client_id == client.id)
    )
    result = await session.execute(stmt)
    candidate = result.scalar_one_or_none()
    if candidate is None:
        raise HTTPException(status_code=404, detail="candidate not found")

    calls_q = await session.execute(
        select(Call).where(Call.candidate_id == candidate_id).order_by(Call.id.desc())
    )
    calls = calls_q.scalars().all()
    return {
        "id": candidate.id,
        "vacancy_id": candidate.vacancy_id,
        "phone": candidate.phone,
        "fio": candidate.fio,
        "source": candidate.source,
        "status": candidate.status,
        "active": candidate.active,
        "created_at": candidate.created_at.isoformat(),
        "calls": [
            {
                "id": c.id,
                "started_at": c.started_at.isoformat() if c.started_at else None,
                "finished_at": c.finished_at.isoformat() if c.finished_at else None,
                "score": c.score,
                "decision": c.decision,
            }
            for c in calls
        ],
    }
